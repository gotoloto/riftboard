#!/usr/bin/env python3
"""
Scrape every tournament deck for a given riftdecks.com champion archetype
and emit decks.json (raw) + cards.json (aggregated for the dashboard).

Usage: python scrape.py [archetype_url]

Defaults to the Lillia, Bashful Bloom Unleashed Constructed page.
"""

import glob
import json
import os
import pathlib
import re
import sys
import time
from datetime import datetime
from urllib.parse import urljoin, urlparse, parse_qs

from bs4 import BeautifulSoup
from curl_cffi import requests

BASE = "https://riftdecks.com"
DEFAULT_URL = (
    "https://riftdecks.com/legends/constructed/lillia-bashful-bloom?metagame_id=3"
)
DECK_HREF_RE = re.compile(r"/riftbound-metagame/deck-")
CARD_HREF_RE = re.compile(r"/cards/details-")


def fetch(url: str, retries: int = 3) -> str:
    last = None
    for attempt in range(retries):
        try:
            r = requests.get(url, impersonate="chrome", timeout=30)
            if r.status_code == 200:
                return r.text
            last = f"status {r.status_code}"
        except Exception as exc:
            last = str(exc)
        time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"fetch failed for {url}: {last}")


RANK_TOP_RE = re.compile(r"Top\s*(\d+)", re.I)
RANK_ORDINAL_RE = re.compile(r"^(\d+)(?:st|nd|rd|th)$", re.I)
PLAYERS_RE = re.compile(r"(\d+)\s*Players", re.I)
DATE_RE = re.compile(r"\b(\d{4}-\d{2}-\d{2})\b")


def _parse_rank(text: str):
    """Convert "Top4", "1st", "76th" -> int rank; return None if unrecognized."""
    text = (text or "").strip()
    m = RANK_TOP_RE.match(text)
    if m:
        return int(m.group(1))
    m = RANK_ORDINAL_RE.match(text)
    if m:
        return int(m.group(1))
    return None


# Names as published on riftdecks that we need to correct site-wide. The
# slug for each remains its riftdecks form (lowercase, no apostrophe), so
# this map only affects user-facing display names.
NAME_FIXES = {
    "Reksai": "Rek'Sai",
}


def normalize_name(s):
    """Apply NAME_FIXES to a card or deck/title string. No-op for non-strings."""
    if not isinstance(s, str):
        return s
    for wrong, right in NAME_FIXES.items():
        s = s.replace(wrong, right)
    return s


def parse_listing(html: str):
    soup = BeautifulSoup(html, "html.parser")
    deck_links = []
    deck_meta: dict[str, dict] = {}
    seen = set()

    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            a = tr.find("a", href=DECK_HREF_RE)
            if not a:
                continue
            full = urljoin(BASE, a["href"]).split("?")[0]
            if full in seen:
                continue
            seen.add(full)
            deck_links.append(full)
            cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
            row_text = " | ".join(cells)
            rank = None
            for c in cells:
                rank = _parse_rank(c)
                if rank is not None:
                    break
            players = None
            m = PLAYERS_RE.search(row_text)
            if m:
                players = int(m.group(1))
            finish_pct = None
            if rank is not None and players and players > 0:
                # Source occasionally reports "Top32" / "Top64" in smaller
                # tournaments (likely a bucket label for a cut that didn't
                # happen). Clamp impossible rank/players ratios at 100%.
                finish_pct = round(min(rank / players, 1.0) * 100, 1)
            d_match = DATE_RE.search(row_text)
            date_str = d_match.group(1) if d_match else None
            deck_meta[full] = {
                "rank": rank,
                "players": players,
                "finish_pct": finish_pct,
                "date": date_str,
            }

    # also find any deck links that weren't in the main tables (fallback)
    for a in soup.find_all("a", href=DECK_HREF_RE):
        full = urljoin(BASE, a["href"]).split("?")[0]
        if full not in seen:
            seen.add(full)
            deck_links.append(full)
            deck_meta.setdefault(
                full,
                {"rank": None, "players": None, "finish_pct": None, "date": None},
            )

    max_page = 1
    for a in soup.find_all("a", href=re.compile(r"page=\d+")):
        m = re.search(r"page=(\d+)", a["href"])
        if m:
            max_page = max(max_page, int(m.group(1)))
    archetype, total = "", None
    if soup.title and soup.title.string:
        title = soup.title.string.strip()
        m = re.match(
            r"Riftbound (?:\w+ )?(.+?) decks - (\d+) available", title
        )
        if m:
            archetype = normalize_name(m.group(1).strip())
            total = int(m.group(2))
        else:
            # Generic search-result title like "Riftbound Top Decks (156 published)"
            m = re.search(r"\((\d+)\s+published\)", title)
            if m:
                total = int(m.group(1))
    return deck_links, max_page, archetype, total, deck_meta


SECTION_RE = re.compile(r"^([\w\s]+?)\s*\((\d+)\)\s*$")
TYPE_NORMALIZE = {
    "battlefields": "battlefield",
    "runes": "rune",
    "spells": "spell",
    "units": "unit",
    "gears": "gear",
    "champions": "champion",
    "legends": "legend",
}


def parse_deck(html: str, url: str):
    soup = BeautifulSoup(html, "html.parser")
    title = soup.title.string.strip() if soup.title and soup.title.string else ""
    h = soup.find(
        lambda t: t.name in ("h2", "h3") and "Text Decklist" in t.get_text()
    )
    if not h:
        return None
    table = h.find_next("table")
    if not table:
        return None
    cards = []
    current_type = None
    board = "main"
    for tr in table.find_all("tr"):
        link = tr.find("a", href=CARD_HREF_RE)
        if not link:
            text = tr.get_text(" ", strip=True)
            m = SECTION_RE.match(text)
            if m:
                section = m.group(1).strip().lower()
                if section == "sideboard":
                    board = "side"
                    current_type = None
                else:
                    current_type = TYPE_NORMALIZE.get(section, section)
            continue
        cell_texts = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        qty = next(
            (int(c) for c in cell_texts if c.isdigit()), None
        )
        if qty is None:
            continue
        href = link["href"]
        slug = href.rsplit("/", 1)[-1]
        cards.append(
            {
                "name": normalize_name(link.get_text(strip=True)),
                "slug": slug,
                "url": urljoin(BASE, href),
                "qty": qty,
                "type": current_type,
                "board": board,
            }
        )
    return {"url": url, "title": normalize_name(title), "cards": cards}


def parse_card_detail(html: str):
    """Returns (fields, image_url). image_url is the first /img/cards/.../_full.png
    on the page, made absolute. Either may be None / empty."""
    soup = BeautifulSoup(html, "html.parser")
    fields: dict[str, list[str]] = {}
    for div in soup.find_all("div", class_="mb-2"):
        h3 = div.find("h3")
        if not h3:
            continue
        key = h3.get_text(strip=True).lower()
        h3.extract()
        val = div.get_text(" ", strip=True)
        if not val:
            continue
        fields.setdefault(key, []).append(val)
    image_url = None
    for img in soup.find_all("img", src=True):
        src = img["src"]
        if "/img/cards/" in src and "_full.png" in src:
            image_url = urljoin(BASE, src.replace("//", "/"))
            break
    return fields, image_url


def save_json(path: str, data) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def save_cards_js(path: str, data) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write("window.__CARDS__ = ")
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write(";\n")
    os.replace(tmp, path)


def build_dashboard_payload(raw: dict) -> dict:
    """Compact, dashboard-friendly view: per-deck mainboard + sideboard summary
    plus card metadata. The dashboard recomputes inclusion/avg/distribution
    stats live from this so UI filters (board, finish percentile, type) can
    change the aggregate."""
    decks_slim = []
    for d in raw["decks"]:
        mainboard: dict[str, int] = {}
        sideboard: dict[str, int] = {}
        for c in d["cards"]:
            target = mainboard if c["board"] == "main" else sideboard
            target[c["slug"]] = target.get(c["slug"], 0) + c["qty"]
        decks_slim.append(
            {
                "u": d["url"],
                "t": d.get("title", ""),
                "rk": d.get("rank"),
                "pl": d.get("players"),
                "fp": d.get("finish_pct"),
                "dt": d.get("date"),
                "c": [[slug, qty] for slug, qty in mainboard.items()],
                "s": [[slug, qty] for slug, qty in sideboard.items()],
            }
        )
    cards_meta = {}
    for slug, m in raw["cards_meta"].items():
        cards_meta[slug] = {
            "name": m.get("name", slug),
            "type": m.get("type"),
            "domains": m.get("domains", []),
            "cost": m.get("cost"),
            "url": m.get("url"),
            "img": m.get("image_url"),
            "rarity": m.get("rarity"),
        }
    return {
        "archetype": raw["archetype"],
        "url": raw["url"],
        "scraped_at": raw["scraped_at"],
        "deck_count": raw["deck_count"],
        "decks": decks_slim,
        "cards_meta": cards_meta,
    }


def save_data_js(path: str, payload: dict) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write("window.__DATA__ = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
    os.replace(tmp, path)


def slug_from_url(url: str) -> str:
    """Extract the archetype slug from a riftdecks URL."""
    m = re.search(r"/constructed/([^/?#]+)", url)
    if m:
        return m.group(1)
    m = re.search(r"/legends/[^/]+/([^/?#]+)", url)
    if m:
        return m.group(1)
    qs = parse_qs(urlparse(url).query)
    if "omni" in qs and qs["omni"]:
        return slugify(qs["omni"][0])
    return ""


def ensure_metagame_param(url: str) -> str:
    """Idempotently append metagame_id=3 to a riftdecks URL so the listing
    stays scoped to the current release period."""
    parsed = urlparse(url)
    qs = parse_qs(parsed.query)
    if qs.get("metagame_id") == ["3"]:
        return url
    sep = "&" if parsed.query else "?"
    return f"{url}{sep}metagame_id=3"


def archetype_from_url(url: str) -> str | None:
    """Fallback archetype name when the page title doesn't include one.

    Title-cases the omni value so a URL like `?omni=azir%2C+emperor` becomes
    `Azir, Emperor` in the dropdown."""
    qs = parse_qs(urlparse(url).query)
    if "omni" in qs and qs["omni"]:
        raw = qs["omni"][0]
        return ", ".join(p.strip().title() for p in raw.split(","))
    return None


def slugify(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "champion"


def legend_dir(slug: str) -> str:
    return os.path.join("legends", slug)


LEGENDS_INDEX_URL = "https://riftdecks.com/legends"
LEGEND_COUNT_RE = re.compile(r"(\d+)\s*Decks?\b", re.I)
LEGEND_SLUG_RE = re.compile(r"/legends/constructed/([a-z0-9-]+)")


def fetch_legends_index() -> dict:
    """Return {slug: current_deck_count} from the /legends index page.

    Used as a cheap freshness probe: rather than walking 17 listing pages per
    archetype to detect new tournament entries, we fetch one page and compare
    counts to what's cached."""
    html = fetch(LEGENDS_INDEX_URL)
    soup = BeautifulSoup(html, "html.parser")
    out: dict = {}
    for a in soup.find_all("a", href=LEGEND_SLUG_RE):
        m = LEGEND_SLUG_RE.search(a["href"])
        if not m:
            continue
        slug = m.group(1)
        if slug in out:
            continue
        container = a.find_parent(["div", "tr", "article", "li", "section"])
        if not container:
            continue
        cm = LEGEND_COUNT_RE.search(container.get_text(" ", strip=True))
        if cm:
            out[slug] = int(cm.group(1))
    return out


def cached_deck_count(slug: str):
    """Return cached deck_count for `slug`, or None if not cached."""
    path = os.path.join(legend_dir(slug), "decks.json")
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f).get("deck_count")
    except FileNotFoundError:
        return None


def list_cached_slugs() -> list:
    return sorted(
        os.path.basename(os.path.dirname(p))
        for p in glob.glob("legends/*/decks.json")
    )


def rebuild_champions_index(directory: str = ".") -> list[dict]:
    """Scan legends/<slug>/data.js and rewrite champions.js."""
    entries = []
    for fn in sorted(glob.glob(os.path.join(directory, "legends", "*", "data.js"))):
        slug = os.path.basename(os.path.dirname(fn))
        try:
            with open(fn, "r", encoding="utf-8") as f:
                txt = f.read()
            start = txt.index("=") + 1
            end = txt.rfind(";")
            data = json.loads(txt[start:end])
        except Exception as exc:
            print(f"      ! could not read {fn}: {exc}")
            continue
        entries.append(
            {
                "slug": slug,
                "name": data.get("archetype", slug),
                "url": data.get("url", ""),
                "scraped_at": data.get("scraped_at", ""),
                "deck_count": data.get("deck_count", 0),
            }
        )
    entries.sort(key=lambda e: e["name"])
    with open(os.path.join(directory, "champions.js"), "w", encoding="utf-8") as f:
        f.write("window.__CHAMPIONS__ = ")
        json.dump(entries, f, ensure_ascii=False, indent=2)
        f.write(";\n")
    return entries


def apply_legend_archetype_label(raw: dict, aggregated: dict) -> None:
    """If a clear legend card exists (>=50% of decks), use its name as the
    archetype label on both the raw and aggregated views, in place. Idempotent
    — safe to call from main() or from a backfill script."""
    legends = sorted(
        (c for c in aggregated["cards"] if (c.get("type") or "").lower() == "legend"),
        key=lambda c: c["decks_including"],
        reverse=True,
    )
    if legends and legends[0]["decks_including"] >= max(1, len(raw["decks"]) // 2):
        raw["archetype"] = legends[0]["name"]
        aggregated["archetype"] = legends[0]["name"]


def aggregate(data: dict) -> dict:
    decks = data["decks"]
    n = len(decks)
    cards_meta = data["cards_meta"]
    # per-deck quantity tally per card
    per_deck_qty: list[dict[str, int]] = []
    seen_slugs: set[str] = set()
    for d in decks:
        totals: dict[str, int] = {}
        for c in d["cards"]:
            if c["board"] != "main":
                continue
            totals[c["slug"]] = totals.get(c["slug"], 0) + c["qty"]
        per_deck_qty.append(totals)
        seen_slugs.update(totals.keys())

    cards = []
    for slug in seen_slugs:
        meta = cards_meta.get(slug, {})
        # build the copies histogram across all n decks (decks not running the
        # card contribute a 0)
        buckets = {0: 0, 1: 0, 2: 0, "3+": 0}
        total_copies = 0
        decks_inc = 0
        for totals in per_deck_qty:
            qty = totals.get(slug, 0)
            if qty == 0:
                buckets[0] += 1
            elif qty == 1:
                buckets[1] += 1
            elif qty == 2:
                buckets[2] += 1
            else:
                buckets["3+"] += 1
            if qty > 0:
                decks_inc += 1
                total_copies += qty
        avg = total_copies / decks_inc if decks_inc else 0
        pct = {k: round(v / n * 100, 1) if n else 0 for k, v in buckets.items()}
        cards.append(
            {
                "slug": slug,
                "name": meta.get("name", slug),
                "type": meta.get("type"),
                "domains": meta.get("domains", []),
                "cost": meta.get("cost"),
                "rarity": meta.get("rarity"),
                "url": meta.get("url"),
                "decks_including": decks_inc,
                "inclusion_pct": round(decks_inc / n * 100, 1) if n else 0,
                "avg_copies_when_included": round(avg, 2),
                "copies_pct": {
                    "0": pct[0],
                    "1": pct[1],
                    "2": pct[2],
                    "3+": pct["3+"],
                },
            }
        )
    cards.sort(key=lambda c: c["decks_including"], reverse=True)
    return {
        "archetype": data["archetype"],
        "url": data["url"],
        "scraped_at": data["scraped_at"],
        "deck_count": n,
        "cards": cards,
    }


def main(archetype_url: str) -> None:
    print(f"[1/4] listing pages — fetching {archetype_url}")
    html = fetch(archetype_url)
    deck_links, max_page, archetype, total, deck_meta = parse_listing(html)
    if not archetype:
        archetype = archetype_from_url(archetype_url) or "Unnamed"
    print(f"      archetype={archetype!r} total={total} pages={max_page}")

    sep = "&" if "?" in archetype_url else "?"
    for p in range(2, max_page + 1):
        url = f"{archetype_url}{sep}page={p}"
        print(f"      page {p}/{max_page}")
        try:
            page_html = fetch(url)
        except Exception as exc:
            print(f"      ! failed: {exc}")
            continue
        more, _, _, _, more_meta = parse_listing(page_html)
        for u in more:
            if u not in deck_links:
                deck_links.append(u)
            if u not in deck_meta and u in more_meta:
                deck_meta[u] = more_meta[u]
        time.sleep(0.4)
    print(f"[1/4] collected {len(deck_links)} unique deck URLs")

    print(f"[2/4] scraping {len(deck_links)} decks")
    interim_slug = slug_from_url(archetype_url) or slugify(archetype)
    interim_dir = legend_dir(interim_slug)
    os.makedirs(interim_dir, exist_ok=True)
    interim_path = os.path.join(interim_dir, "decks.json")
    decks: list[dict] = []
    for i, durl in enumerate(deck_links, 1):
        try:
            page_html = fetch(durl)
            d = parse_deck(page_html, durl)
            if d:
                meta = deck_meta.get(durl, {})
                d["rank"] = meta.get("rank")
                d["players"] = meta.get("players")
                d["finish_pct"] = meta.get("finish_pct")
                d["date"] = meta.get("date")
                decks.append(d)
            else:
                print(f"      ! no decklist parsed: {durl}")
        except Exception as exc:
            print(f"      ! deck {i} failed: {exc}")
        if i % 10 == 0:
            print(f"      {i}/{len(deck_links)} ({len(decks)} parsed)")
            save_json(
                interim_path,
                {
                    "archetype": archetype,
                    "url": archetype_url,
                    "scraped_at": datetime.utcnow().isoformat() + "Z",
                    "deck_count": len(decks),
                    "decks": decks,
                    "cards_meta": {},
                },
            )
        time.sleep(0.3)
    print(f"[2/4] parsed {len(decks)} decks")

    print("[3/4] fetching unique-card metadata (type/domains/cost)")
    unique: dict[str, dict] = {}
    for d in decks:
        for c in d["cards"]:
            if c["slug"] not in unique:
                unique[c["slug"]] = {
                    "name": c["name"],
                    "url": c["url"],
                    "type": c["type"],
                }
    # Catalog-aware shortcut: any slug already in cards-catalog.json has its
    # canonical printing metadata recorded — we skip the detail-page fetch
    # and copy the fields verbatim. New slugs still get fetched live.
    catalog = load_catalog()
    cached_hits = 0
    for i, (slug, info) in enumerate(unique.items(), 1):
        if enrich_from_catalog(info, catalog.get(slug)):
            cached_hits += 1
            continue
        try:
            page_html = fetch(info["url"])
            fields, image_url = parse_card_detail(page_html)
            info["domains"] = fields.get("domains", [])
            cost_vals = fields.get("cost", [])
            info["cost"] = cost_vals[0] if cost_vals else None
            type_vals = fields.get("types", [])
            if type_vals:
                info["type"] = type_vals[0]
            rarities = fields.get("rarity") or []
            non_showcase = [r for r in rarities if (r or "").lower() != "showcase"]
            info["rarity"] = (non_showcase or rarities or [None])[0]
            info["image_url"] = image_url
        except Exception as exc:
            print(f"      ! card {slug} failed: {exc}")
        if i % 25 == 0:
            print(f"      card {i}/{len(unique)}")
        time.sleep(0.25)
    print(
        f"[3/4] enriched {len(unique)} cards "
        f"({cached_hits} from catalog, {len(unique) - cached_hits} freshly fetched)"
    )

    raw = {
        "archetype": archetype,
        "url": archetype_url,
        "scraped_at": datetime.utcnow().isoformat() + "Z",
        "deck_count": len(decks),
        "decks": decks,
        "cards_meta": unique,
    }
    slug = slug_from_url(archetype_url) or slugify(archetype)
    out_dir = legend_dir(slug)
    os.makedirs(out_dir, exist_ok=True)
    decks_path = os.path.join(out_dir, "decks.json")
    data_path = os.path.join(out_dir, "data.js")

    print(f"[4/4] aggregating {data_path}")
    aggregated = aggregate(raw)
    # Prefer the actual legend card's name as the archetype label.
    apply_legend_archetype_label(raw, aggregated)
    save_json(decks_path, raw)
    save_data_js(data_path, build_dashboard_payload(raw))
    rebuild_champions_index()
    print(f"      {len(aggregated['cards'])} unique cards over {raw['deck_count']} decks")
    print(f"      slug: {slug}")
    print("done.")


def update_archetype(slug: str) -> dict:
    """Incremental update for a cached champion. Re-walks the listing pages
    (to pick up new tournament entries and refresh rank/players/date on
    existing ones) and fetches only the deck pages whose URL isn't already
    cached. Card metadata is fetched only for unseen slugs."""
    out_dir = legend_dir(slug)
    decks_path = os.path.join(out_dir, "decks.json")
    raw = json.load(open(decks_path, encoding="utf-8"))
    archetype_url = ensure_metagame_param(raw["url"])
    raw["url"] = archetype_url

    print(f"  listing {archetype_url}")
    html = fetch(archetype_url)
    deck_links, max_page, _, _, deck_meta = parse_listing(html)
    sep = "&" if "?" in archetype_url else "?"
    for p in range(2, max_page + 1):
        try:
            page_html = fetch(f"{archetype_url}{sep}page={p}")
        except Exception as exc:
            print(f"    ! page {p}: {exc}")
            continue
        more, _, _, _, more_meta = parse_listing(page_html)
        for u in more:
            if u not in deck_meta and u in more_meta:
                deck_meta[u] = more_meta[u]
            if u not in deck_links:
                deck_links.append(u)
        time.sleep(0.4)
    print(f"  listed {len(deck_links)} decks across {max_page} pages")

    cached_by_url = {d["url"]: d for d in raw["decks"]}
    new_urls = [u for u in deck_links if u not in cached_by_url]
    print(f"  cached={len(cached_by_url)}, new={len(new_urls)}")

    refreshed = 0
    for u, m in deck_meta.items():
        d = cached_by_url.get(u)
        if not d or not m:
            continue
        before = (d.get("rank"), d.get("players"), d.get("finish_pct"), d.get("date"))
        after = (m.get("rank"), m.get("players"), m.get("finish_pct"), m.get("date"))
        if before != after:
            d["rank"], d["players"], d["finish_pct"], d["date"] = after
            refreshed += 1

    new_decks = []
    for i, durl in enumerate(new_urls, 1):
        try:
            page_html = fetch(durl)
            d = parse_deck(page_html, durl)
        except Exception as exc:
            print(f"    ! new deck {i}: {exc}")
            continue
        if not d:
            continue
        m = deck_meta.get(durl, {})
        d["rank"] = m.get("rank")
        d["players"] = m.get("players")
        d["finish_pct"] = m.get("finish_pct")
        d["date"] = m.get("date")
        new_decks.append(d)
        if i % 10 == 0:
            print(f"    new deck {i}/{len(new_urls)}")
        time.sleep(0.3)
    raw["decks"].extend(new_decks)
    raw["deck_count"] = len(raw["decks"])

    cards_meta = raw["cards_meta"]
    new_slugs = []
    for d in new_decks:
        for c in d["cards"]:
            if c["slug"] in cards_meta:
                continue
            cards_meta[c["slug"]] = {
                "name": c["name"],
                "url": c["url"],
                "type": c["type"],
            }
            new_slugs.append(c["slug"])
    # Catalog-aware shortcut for newly-seen slugs.
    catalog = load_catalog()
    cached_hits = 0
    for slug2 in new_slugs:
        info = cards_meta[slug2]
        if enrich_from_catalog(info, catalog.get(slug2)):
            cached_hits += 1
            continue
        try:
            fields, image_url = parse_card_detail(fetch(info["url"]))
        except Exception as exc:
            print(f"    ! card {slug2}: {exc}")
            continue
        info["domains"] = fields.get("domains", [])
        cost_vals = fields.get("cost", [])
        info["cost"] = cost_vals[0] if cost_vals else None
        type_vals = fields.get("types", [])
        if type_vals:
            info["type"] = type_vals[0]
        rarities = fields.get("rarity") or []
        non_showcase = [r for r in rarities if (r or "").lower() != "showcase"]
        info["rarity"] = (non_showcase or rarities or [None])[0]
        info["image_url"] = image_url
        time.sleep(0.25)
    if new_slugs:
        print(
            f"    {len(new_slugs)} new slug(s): {cached_hits} from catalog, "
            f"{len(new_slugs) - cached_hits} freshly fetched"
        )

    raw["scraped_at"] = datetime.utcnow().isoformat() + "Z"
    aggregated = aggregate(raw)
    apply_legend_archetype_label(raw, aggregated)
    save_json(decks_path, raw)
    save_data_js(os.path.join(out_dir, "data.js"), build_dashboard_payload(raw))
    return {
        "new_decks": len(new_decks),
        "refreshed_meta": refreshed,
        "new_cards": len(new_slugs),
        "total_decks": raw["deck_count"],
    }


def build_staples(top_per_rarity: int = 40) -> dict:
    """Aggregate top cards by total decks_including across every cached
    legend, bucketed by rarity, excluding runes and battlefields. For each
    card we record which legends play it in >50% of their decks.

    Reads decks.json directly and computes per-legend stats in memory via
    aggregate() — the previous version read pre-computed cards.json files
    but those are derived from the same source, so eliminating them removes
    drift risk without changing the output."""
    legend_info: dict = {}
    image_index: dict = {}
    agg: dict = {}

    for decks_path in sorted(glob.glob("legends/*/decks.json")):
        legend_slug = os.path.basename(os.path.dirname(decks_path))
        try:
            raw = json.load(open(decks_path, encoding="utf-8"))
        except Exception:
            continue
        legend_info[legend_slug] = {
            "name": raw.get("archetype", legend_slug),
            "deck_count": raw.get("deck_count", 0),
        }
        # Per-card stats (same shape cards.json used to carry).
        per_legend = aggregate(raw)
        for c in per_legend.get("cards", []):
            slug = c.get("slug")
            if not slug:
                continue
            t = (c.get("type") or "").lower()
            if t in ("rune", "battlefield"):
                continue
            rarity = (c.get("rarity") or "").lower()
            entry = agg.setdefault(
                slug,
                {
                    "slug": slug,
                    "name": c.get("name", slug),
                    "type": t,
                    "rarity": rarity,
                    "domains": c.get("domains") or [],
                    "url": c.get("url"),
                    "total_decks_including": 0,
                    "_per_legend": {},
                },
            )
            if not entry["domains"] and c.get("domains"):
                entry["domains"] = c["domains"]
            if not entry["rarity"] and rarity:
                entry["rarity"] = rarity
            entry["total_decks_including"] += c.get("decks_including", 0)
            entry["_per_legend"][legend_slug] = {
                "decks_including": c.get("decks_including", 0),
                "inclusion_pct": c.get("inclusion_pct", 0),
            }
        # Image index from cards_meta on the same decks.json.
        for slug, m in raw.get("cards_meta", {}).items():
            if slug not in image_index and m.get("image_url"):
                image_index[slug] = m["image_url"]

    for entry in agg.values():
        if entry["slug"] in image_index:
            entry["img"] = image_index[entry["slug"]]
        legends_above = [
            {
                "slug": ls,
                "name": legend_info.get(ls, {}).get("name", ls),
                "inclusion_pct": round(pl["inclusion_pct"] or 0, 1),
                "decks_including": pl["decks_including"],
            }
            for ls, pl in entry["_per_legend"].items()
            if (pl["inclusion_pct"] or 0) > 50
        ]
        legends_above.sort(key=lambda x: -x["inclusion_pct"])
        entry["legends_above_50pct"] = legends_above
        del entry["_per_legend"]

    by_rarity = {"common": [], "uncommon": [], "rare": []}
    for entry in agg.values():
        if entry["rarity"] in by_rarity:
            by_rarity[entry["rarity"]].append(entry)
    for r, lst in by_rarity.items():
        lst.sort(key=lambda x: (-x["total_decks_including"], x["name"]))
        by_rarity[r] = lst[:top_per_rarity]

    return {
        "scraped_at": datetime.utcnow().isoformat() + "Z",
        "total_decks": sum(L["deck_count"] for L in legend_info.values()),
        "total_legends": len(legend_info),
        "top_per_rarity": top_per_rarity,
        "rarities": by_rarity,
    }


CATALOG_PATH = "cards-catalog.json"
PRINTING_RE = re.compile(
    r"/img/cards/[^/]+/+([A-Z][A-Z0-9]+)/[a-z][a-z0-9]+-(\d+)([a-z]*)-(\d+)_full\.png"
)
# Riftdecks embeds the TCGplayer market price in the meta description of
# each card page: e.g. <meta name="description" content="...around $2.58.">
# We pull from there because the in-body chip's markup changes more often
# than the meta tag does.
PRICE_RE = re.compile(
    r'<meta[^>]*name="description"[^>]*content="[^"]*?around \$([0-9]+(?:\.[0-9]+)?)',
    re.IGNORECASE,
)


def extract_card_price(html: str):
    """Pull TCGplayer market price (USD) from a card detail page's meta
    description. Returns float or None if not present."""
    m = PRICE_RE.search(html)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def fetch_card_catalog(slugs=None) -> dict:
    """Walk /cards for every card slug, fetch each detail page, and record
    the *canonical* printing (lowest numeric collector # that's in-range and
    has no alt-art letter suffix). Overnumbered-only cards are dropped."""
    if slugs is None:
        html = fetch("https://riftdecks.com/cards")
        soup = BeautifulSoup(html, "html.parser")
        slugs = sorted(
            {
                a["href"].rsplit("/", 1)[-1]
                for a in soup.find_all("a", href=re.compile(r"/cards/details-"))
            }
        )
        print(f"  {len(slugs)} slugs on /cards")
    out: dict = {}
    skipped = 0
    for i, slug in enumerate(slugs, 1):
        url = urljoin(BASE, f"/cards/{slug}")
        try:
            page_html = fetch(url)
        except Exception as exc:
            print(f"  ! {slug}: {exc}")
            continue
        fields, _ = parse_card_detail(page_html)
        soup = BeautifulSoup(page_html, "html.parser")
        printings = []
        for img in soup.find_all("img", src=True):
            m = PRINTING_RE.match(img["src"])
            if m:
                printings.append(
                    {
                        "set": m.group(1),
                        "num": int(m.group(2)),
                        "suffix": m.group(3),
                        "setmax": int(m.group(4)),
                        "src": img["src"],
                    }
                )
        # Canonical: no letter suffix AND num within the set's standard range.
        canon = sorted(
            (p for p in printings if not p["suffix"] and p["num"] <= p["setmax"]),
            key=lambda p: p["num"],
        )
        if not canon:
            skipped += 1
            continue  # Overnumbered-only / alt-art-only — drop per user spec.
        c = canon[0]
        # The detail page lists "Rarity" twice: once for the currently-displayed
        # printing and once for the canonical card. Showcase entries indicate
        # alt-art premium printings — never what a user actually buys for play.
        rarities = fields.get("rarity") or []
        non_showcase = [r for r in rarities if (r or "").lower() != "showcase"]
        rarity = (non_showcase or rarities or [""])[0].lower()

        # Build a list of every printing URL: the main URL plus all variant
        # links of the form /cards/<slug>/<N>. Then pick the cheapest one
        # whose own rarity list contains NO 'showcase' entry. This handles,
        # uniformly:
        #   - Showcase variants (alt-art, often 10-1000x the standard).
        #   - Holographic/foil variants (sometimes higher than standard,
        #     sometimes lower depending on print runs).
        #   - Cards where the main URL happens to default to a premium
        #     printing (e.g. Kai'sa, Daughter of the Void — main shows the
        #     showcase chase at $2331, but /186 is the $0.31 standard).
        # By taking the minimum non-showcase price, we always converge on
        # the cheapest playable copy the user could actually buy.
        variant_urls = []
        for a in soup.find_all("a", href=True):
            vm = re.match(
                r"^/cards/" + re.escape(slug) + r"/\d+$",
                a["href"],
            )
            if vm and a["href"] not in variant_urls:
                variant_urls.append(a["href"])
        # Riftdecks' rarity list has *two* entries per page: position 0 is
        # the rarity of the currently-DISPLAYED printing, position 1 is some
        # canonical-card metadata that's identical across all variants of
        # the same slug. So to decide whether a page is showing a showcase
        # printing we look only at the first entry. Position 1 is noise for
        # this decision.
        def displayed_rarity(rar_list):
            return (rar_list[0] or "").lower() if rar_list else ""

        # Candidate set: each entry is (price, html, rarity_list, source).
        candidates = []
        if displayed_rarity(rarities) != "showcase":
            p = extract_card_price(page_html)
            if p is not None:
                candidates.append((p, page_html, rarities, "main"))
        for vurl in variant_urls:
            try:
                vhtml = fetch(urljoin(BASE, vurl))
            except Exception:
                continue
            vfields, _ = parse_card_detail(vhtml)
            vrar = vfields.get("rarity") or []
            if displayed_rarity(vrar) == "showcase":
                time.sleep(0.15)
                continue
            p = extract_card_price(vhtml)
            if p is not None:
                candidates.append((p, vhtml, vrar, vurl))
            time.sleep(0.15)
        # Choose the cheapest non-showcase candidate. If the main page rarity
        # was showcase-only, also adopt rarity from the winning variant.
        price = None
        if candidates:
            candidates.sort(key=lambda x: x[0])
            price, _winner_html, winner_rar, _src = candidates[0]
            if rarity == "showcase":
                winner_non = [r for r in winner_rar if (r or "").lower() != "showcase"]
                if winner_non:
                    rarity = winner_non[0].lower()
        out[slug] = {
            "slug": slug,
            "name": normalize_name((fields.get("name") or [slug])[0]),
            "type": (fields.get("types") or [""])[0].lower(),
            "domains": fields.get("domains", []),
            "cost": (fields.get("cost") or [None])[0],
            "rarity": rarity,
            "set": c["set"],
            "set_num": c["num"],
            "set_max": c["setmax"],
            "image_url": urljoin(BASE, c["src"].replace("//", "/")),
            "url": url,
            "price": price,
            # Number of distinct printing URLs (main + variants) — useful for
            # spotting cards with foil/showcase/holo variants so future runs
            # can target multi-printing slugs without re-discovering them.
            "printing_count": 1 + len(variant_urls),
        }
        if i % 25 == 0:
            print(f"    {i}/{len(slugs)} (kept {len(out)}, skipped {skipped})")
        time.sleep(0.2)
    print(f"  catalog: {len(out)} canonical, {skipped} skipped (no in-range printing)")
    return out


def import_collection_xlsx(
    xlsx_path: str,
    owned_path: str = "collection-owned.js",
    enroute_path: str = "collection-enroute.js",
) -> dict:
    """Read a filled-in collection xlsx and emit JS files mapping
    slug → qty owned (and slug → qty en route, if the column exists).
    The cart page reads `__OWNED_DEFAULTS__` as the Owned baseline; the
    closeness page can toggle in `__EN_ROUTE_DEFAULTS__` so cards
    in-transit count toward completeness."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("openpyxl is required: pip install openpyxl")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    try:
        slug_col = headers.index("Slug")
        owned_col = headers.index("Qty Owned")
    except ValueError as exc:
        raise SystemExit(f"missing required column: {exc}")
    enroute_col = headers.index("Qty En Route") if "Qty En Route" in headers else None

    def parse_qty(v):
        if v is None or v == "":
            return None
        try:
            n = int(v)
        except (TypeError, ValueError):
            return None
        return n if n > 0 else None

    owned: dict = {}
    enroute: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        slug = row[slug_col]
        if not slug:
            continue
        s = str(slug).strip()
        oq = parse_qty(row[owned_col])
        if oq is not None:
            owned[s] = oq
        if enroute_col is not None:
            eq = parse_qty(row[enroute_col])
            if eq is not None:
                enroute[s] = eq

    def write_js(path, varname, data):
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(f"window.{varname} = ")
            json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
            f.write(";\n")
        os.replace(tmp, path)

    write_js(owned_path, "__OWNED_DEFAULTS__", owned)
    write_js(enroute_path, "__EN_ROUTE_DEFAULTS__", enroute)
    return {
        "owned_path": owned_path,
        "owned_distinct": len(owned),
        "owned_total": sum(owned.values()),
        "enroute_path": enroute_path,
        "enroute_distinct": len(enroute),
        "enroute_total": sum(enroute.values()),
        "enroute_column_present": enroute_col is not None,
    }


DECK_URL_PREFIX = "https://riftdecks.com/riftbound-metagame/"


def build_deck_lookup_data(output_path: str = "deck-lookup.js") -> dict:
    """Flat index keyed by deck URL → that deck's cards + tournament metadata.
    Used by /diff.html so the user can paste a riftdecks deck URL and we
    pick up the cards from local cache instead of fighting CORS/Cloudflare.

    Slugs are interned into a single string table to keep the payload small
    (every deck reuses the same ~500 card slugs)."""
    slug_index: dict = {}
    slugs_list: list = []

    def sid(slug):
        if slug not in slug_index:
            slug_index[slug] = len(slugs_list)
            slugs_list.append(slug)
        return slug_index[slug]

    decks_out: dict = {}
    for decks_path in sorted(glob.glob("legends/*/decks.json")):
        legend_slug = os.path.basename(os.path.dirname(decks_path))
        try:
            raw = json.load(open(decks_path, encoding="utf-8"))
        except Exception:
            continue
        legend_name = raw.get("archetype", legend_slug)
        for d in raw.get("decks", []):
            url = d.get("url") or ""
            if not url:
                continue
            key = url[len(DECK_URL_PREFIX):] if url.startswith(DECK_URL_PREFIX) else url
            main: dict = {}
            side: dict = {}
            for c in d.get("cards", []):
                m = main if c.get("board") == "main" else side
                m[c["slug"]] = m.get(c["slug"], 0) + c.get("qty", 0)
            decks_out[key] = {
                "t": d.get("title", ""),
                "ln": legend_name,
                "rk": d.get("rank"),
                "pl": d.get("players"),
                "fp": d.get("finish_pct"),
                "dt": d.get("date"),
                "c": [[sid(s), q] for s, q in main.items()],
                "s": [[sid(s), q] for s, q in side.items()],
            }

    payload = {
        "scraped_at": datetime.utcnow().isoformat() + "Z",
        "url_prefix": DECK_URL_PREFIX,
        "slugs": slugs_list,
        "deck_count": len(decks_out),
        "decks": decks_out,
    }
    tmp = output_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write("window.__DECK_LOOKUP__ = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
    os.replace(tmp, output_path)
    return {
        "path": output_path,
        "deck_count": len(decks_out),
        "slugs": len(slugs_list),
    }


def save_catalog_js(catalog: dict, path: str = "cards-catalog.js") -> None:
    """JS-wrapper twin of cards-catalog.json so dashboard pages can load the
    catalog with a plain <script> tag (works under file:// and avoids extra
    fetch() round-trips)."""
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write("window.__CATALOG__ = ")
        json.dump(catalog, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
    os.replace(tmp, path)


def save_catalog_json(catalog: dict, path: str = CATALOG_PATH) -> None:
    save_json(
        path,
        {
            "scraped_at": datetime.utcnow().isoformat() + "Z",
            "card_count": len(catalog),
            "cards": catalog,
        },
    )


def load_catalog() -> dict:
    try:
        with open(CATALOG_PATH, encoding="utf-8") as f:
            return json.load(f).get("cards", {}) or {}
    except FileNotFoundError:
        return {}


def enrich_from_catalog(info: dict, catalog_entry) -> bool:
    """Copy domains/cost/type/rarity/image_url from a cards-catalog.json entry
    into a per-legend cards_meta entry, so the deck scrape can skip the
    expensive card-detail-page fetch. Returns True on hit, False if the
    catalog has no entry for this slug.

    Only fills fields the caller didn't already provide (so we don't stomp
    on values that came from the deck page itself, e.g. the name)."""
    if not catalog_entry:
        return False
    info.setdefault("name", catalog_entry.get("name"))
    info.setdefault("url", catalog_entry.get("url"))
    info["type"] = catalog_entry.get("type") or info.get("type")
    info["domains"] = catalog_entry.get("domains") or []
    info["cost"] = catalog_entry.get("cost")
    info["rarity"] = catalog_entry.get("rarity")
    info["image_url"] = catalog_entry.get("image_url")
    if catalog_entry.get("price") is not None:
        info["price"] = catalog_entry["price"]
    return True


def find_uncatalogued_slugs() -> list:
    """Slugs referenced by any cached legend's decks.json (via its
    cards_meta — which is the slug index inside decks.json) but missing
    from cards-catalog.json. Used by --catalog-new to backfill only the
    gap instead of re-walking all ~770 card detail pages.

    Note: only finds slugs we've already seen via deck scrapes. A brand-
    new card that hasn't appeared in any tournament deck yet won't show
    up here — for that you still want full --catalog (which discovers via
    /cards). In practice the gap closes quickly because new sets get
    play-tested fast."""
    catalog = load_catalog()
    known = set(catalog.keys())
    found: set = set()
    legends_dir = pathlib.Path("legends")
    if not legends_dir.exists():
        return []
    for d in sorted(legends_dir.iterdir()):
        if not d.is_dir():
            continue
        decks_path = d / "decks.json"
        if not decks_path.exists():
            continue
        try:
            data = json.loads(decks_path.read_text())
        except Exception:
            continue
        cards_meta = data.get("cards_meta") or {}
        found.update(cards_meta.keys())
    return sorted(found - known)


def build_collection_template(path: str = "collection-template.xlsx") -> dict:
    """Walk every cached legend's cards_meta, dedupe across legends by slug,
    and emit an Excel template the user can fill in to record what they own.
    The cart page can read this file back to auto-populate the Owned column."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("openpyxl is required: pip install openpyxl")

    catalog = load_catalog()
    # Legend cards on the source's detail page only show the epithet
    # (e.g. "Scorn of the Moon" instead of "Diana, Scorn of the Moon").
    # Restore the full archetype identity from champions.js for those rows.
    legend_full_names: dict = {}
    try:
        with open("champions.js", encoding="utf-8") as f:
            txt = f.read()
        start = txt.index("=") + 1
        end = txt.rfind(";")
        for entry in json.loads(txt[start:end]):
            legend_full_names[f"details-{entry['slug']}"] = entry["name"]
    except Exception:
        pass

    if catalog:
        rows = sorted(
            (
                {
                    "slug": slug,
                    "name": legend_full_names.get(slug)
                    if (m.get("type") or "").lower() == "legend"
                    and legend_full_names.get(slug)
                    else m.get("name", slug),
                    "set": m.get("set", ""),
                    "set_num_raw": str(m.get("set_num") or ""),
                    "set_num_int": m.get("set_num") or 10**9,
                    "domains": ", ".join(m.get("domains") or []),
                    "rarity": (m.get("rarity") or "").lower(),
                    "type": (m.get("type") or "").lower(),
                }
                for slug, m in catalog.items()
            ),
            key=lambda r: (
                r["set"] or "￿",
                r["set_num_int"],
                r["name"].lower(),
            ),
        )
    else:
        # Fallback: per-legend cards_meta (older, less accurate — has only the
        # printings we happened to scrape; may include overnumbered-only cards
        # and a/b suffixes). Run `scrape.py --catalog` to populate the catalog.
        seen: dict = {}
        for decks_path in sorted(glob.glob("legends/*/decks.json")):
            try:
                raw = json.load(open(decks_path, encoding="utf-8"))
            except Exception:
                continue
            for slug, m in raw.get("cards_meta", {}).items():
                if slug in seen:
                    continue
                img = m.get("image_url") or ""
                m_url = PRINTING_RE.match(img.replace("https://riftdecks.com", ""))
                if not m_url:
                    m_url = PRINTING_RE.match(img)
                seen[slug] = {
                    "slug": slug,
                    "name": m.get("name", slug),
                    "set": m_url.group(1) if m_url else "",
                    "set_num_raw": m_url.group(2) if m_url else "",  # numeric only
                    "set_num_int": int(m_url.group(2)) if m_url else 10**9,
                    "domains": ", ".join(m.get("domains", []) or []),
                    "rarity": (m.get("rarity") or "").lower(),
                    "type": (m.get("type") or "").lower(),
                }
        rows = sorted(
            seen.values(),
            key=lambda r: (
                r["set"] or "￿",
                r["set_num_int"],
                r["name"].lower(),
            ),
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"

    headers = [
        "Slug",
        "Card",
        "Set",
        "Set #",
        "Domains",
        "Rarity",
        "Type",
        "Qty Owned",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    for r in rows:
        ws.append([
            r["slug"],
            r["name"],
            r["set"],
            r["set_num_raw"],
            r["domains"],
            r["rarity"],
            r["type"],
            "",  # Qty Owned — user fills this in
        ])

    last_row = ws.max_row
    last_col = len(headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "B2"

    widths = {
        "A": 36,  # Slug
        "B": 32,  # Card
        "C": 8,   # Set
        "D": 8,   # Set #
        "E": 18,  # Domains
        "F": 12,  # Rarity
        "G": 12,  # Type
        "H": 12,  # Qty Owned
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # Right-align Qty Owned
    qty_col = get_column_letter(len(headers))
    for row in range(2, last_row + 1):
        ws[f"{qty_col}{row}"].alignment = Alignment(horizontal="right")

    wb.save(path)
    return {"path": path, "rows": last_row - 1}


def build_closeness_data(
    percentile: float = 25.0,
    output_path: str = "closeness-data.js",
) -> dict:
    """For each cached legend, build the top-percentile *composite* deck —
    1 legend + 3 battlefields + 12 runes + 40 maindeck cards using a
    sort-by-inclusion + greedy-median pick per slot (same shape as the
    main dashboard's Composite mode). Saved as closeness-data.js so the
    /closeness.html page can score it against the user's collection
    without lazy-loading every legend's data.js."""
    static_targets = {"legend": 1, "battlefield": 3, "rune": 12}
    maindeck_target = 40
    excluded_from_main = {"legend", "battlefield", "rune"}

    def pick_by_median(pool, target):
        picks = []
        filled = 0
        for c in pool:
            if filled >= target:
                break
            wanted = max(1, c["median_copies"])
            copies = min(wanted, target - filled)
            picks.append({**c, "qty": copies})
            filled += copies
        return picks

    out_legends = []
    for decks_path in sorted(glob.glob("legends/*/decks.json")):
        slug = os.path.basename(os.path.dirname(decks_path))
        try:
            raw = json.load(open(decks_path, encoding="utf-8"))
        except Exception:
            continue
        decks = [
            d
            for d in raw.get("decks", [])
            if d.get("finish_pct") is not None and d["finish_pct"] <= percentile
        ]
        if not decks:
            continue
        cards_meta = raw.get("cards_meta", {})
        n = len(decks)
        per_card: dict = {}
        for d in decks:
            for c in d.get("cards", []):
                if c.get("board") != "main":
                    continue
                cs = c["slug"]
                per_card.setdefault(cs, []).append(c.get("qty", 0))
        cards = []
        for cs, qtys in per_card.items():
            qtys.sort()
            mid = len(qtys) // 2
            med = qtys[mid] if len(qtys) % 2 else (qtys[mid - 1] + qtys[mid]) / 2
            meta = cards_meta.get(cs, {})
            cards.append(
                {
                    "slug": cs,
                    "name": meta.get("name", cs),
                    "type": (meta.get("type") or "").lower(),
                    "rarity": (meta.get("rarity") or "").lower(),
                    "decks_including": len(qtys),
                    "median_copies": int(round(med)),
                }
            )
        cards.sort(
            key=lambda c: (
                -c["decks_including"],
                -c["median_copies"],
                c["name"],
            )
        )
        composite = []
        for key in ("legend", "battlefield", "rune"):
            composite += pick_by_median(
                [c for c in cards if c["type"] == key],
                static_targets[key],
            )
        composite += pick_by_median(
            [c for c in cards if c["type"] not in excluded_from_main],
            maindeck_target,
        )
        out_legends.append(
            {
                "slug": slug,
                "name": raw.get("archetype", slug),
                "filtered_deck_count": n,
                "composite": [
                    {
                        "slug": c["slug"],
                        "name": c["name"],
                        "qty": c["qty"],
                        "rarity": c["rarity"],
                        "type": c["type"],
                    }
                    for c in composite
                ],
            }
        )
    out_legends.sort(key=lambda L: L["name"])
    payload = {
        "scraped_at": datetime.utcnow().isoformat() + "Z",
        "percentile": percentile,
        "legends": out_legends,
    }
    tmp = output_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write("window.__CLOSENESS_DATA__ = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
    os.replace(tmp, output_path)
    return {"path": output_path, "legends": len(out_legends)}


def save_staples_js(payload: dict, path: str = "staples.js") -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        f.write("window.__STAPLES__ = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
    os.replace(tmp, path)


if __name__ == "__main__":
    args = sys.argv[1:]
    if args and args[0] == "--staples":
        payload = build_staples()
        save_staples_js(payload)
        n = sum(len(v) for v in payload["rarities"].values())
        print(f"staples.js written: {n} cards across {payload['total_decks']} decks / {payload['total_legends']} legends")
    elif args and args[0] == "--catalog":
        catalog = fetch_card_catalog()
        save_catalog_json(catalog)
        save_catalog_js(catalog)
        print(f"{CATALOG_PATH} + cards-catalog.js written: {len(catalog)} cards")
    elif args and args[0] == "--catalog-new":
        # Incremental: fetch detail pages only for slugs referenced by some
        # legend's cards.json but missing from cards-catalog.json. Avoids
        # re-walking all ~770 cards when only a handful are new.
        existing = load_catalog()
        missing = find_uncatalogued_slugs()
        if not missing:
            print(f"catalog already has all {len(existing)} referenced cards · nothing to do")
        else:
            print(f"existing catalog: {len(existing)} cards · fetching {len(missing)} new slug(s):")
            for s in missing:
                print(f"  + {s}")
            new_entries = fetch_card_catalog(slugs=missing)
            merged = {**existing, **new_entries}
            save_catalog_json(merged)
            save_catalog_js(merged)
            print(f"{CATALOG_PATH} + cards-catalog.js written: {len(merged)} cards (+{len(merged) - len(existing)} new)")
    elif args and args[0] == "--deck-lookup":
        info = build_deck_lookup_data()
        print(f"{info['path']} written: {info['deck_count']} decks")
    elif args and args[0] == "--closeness":
        info = build_closeness_data()
        print(f"{info['path']} written: {info['legends']} legends")
    elif args and args[0] == "--collection":
        info = build_collection_template()
        print(f"{info['path']} written: {info['rows']} unique cards")
    elif args and args[0] == "--import-collection":
        if len(args) < 2:
            raise SystemExit("usage: scrape.py --import-collection <path-to-xlsx>")
        info = import_collection_xlsx(args[1])
        print(
            f"{info['owned_path']} written: "
            f"{info['owned_distinct']} distinct, {info['owned_total']} copies"
        )
        if info["enroute_column_present"]:
            print(
                f"{info['enroute_path']} written: "
                f"{info['enroute_distinct']} distinct, "
                f"{info['enroute_total']} copies"
            )
        else:
            print(
                f"  (no 'Qty En Route' column in xlsx — wrote empty {info['enroute_path']})"
            )
    elif args and args[0] == "--check":
        print(f"Pinging {LEGENDS_INDEX_URL}…")
        index = fetch_legends_index()
        cached = list_cached_slugs()
        seen = set()
        for s in cached:
            seen.add(s)
            local = cached_deck_count(s)
            current = index.get(s)
            if current is None:
                print(f"  ? {s}: cached={local}, not on /legends")
            elif current == local:
                print(f"  = {s}: {local}")
            else:
                delta = current - (local or 0)
                sign = "+" if delta >= 0 else ""
                print(f"  → {s}: {local} → {current}  ({sign}{delta})")
        new = sorted((s, c) for s, c in index.items() if s not in seen)
        for s, c in new:
            print(f"  + {s}: {c} (not cached — run `scrape.py {LEGENDS_INDEX_URL[:-len('/legends')]}/legends/constructed/{s}?metagame_id=3` to add)")
    elif args and args[0] == "--update":
        explicit_slugs = args[1:]
        if explicit_slugs:
            slugs = explicit_slugs
        else:
            print(f"Checking {LEGENDS_INDEX_URL} for changes…")
            index = fetch_legends_index()
            slugs = []
            for s in list_cached_slugs():
                local = cached_deck_count(s)
                current = index.get(s)
                if current is None:
                    print(f"  ? {s}: not on /legends (skip)")
                    continue
                if current == local:
                    print(f"  = {s}: {local} (unchanged, skip)")
                    continue
                delta = current - (local or 0)
                sign = "+" if delta >= 0 else ""
                print(f"  → {s}: {local} → {current} ({sign}{delta}) (update)")
                slugs.append(s)
        for s in slugs:
            print(f"=== {s} ===")
            try:
                stats = update_archetype(s)
                print(f"  {stats}")
            except FileNotFoundError:
                print(f"  ! legends/{s}/decks.json not found; skip")
        rebuild_champions_index()
        save_staples_js(build_staples())
        build_collection_template()
        build_closeness_data()
        build_deck_lookup_data()
        # Catalog js is a wrapper for the existing JSON; cheap to regen.
        try:
            save_catalog_js(load_catalog())
        except Exception:
            pass
        print(
            "staples.js + collection-template.xlsx + closeness-data.js + deck-lookup.js refreshed"
        )
    else:
        url = args[0] if args else DEFAULT_URL
        main(url)
        save_staples_js(build_staples())
        build_collection_template()
        build_closeness_data()
        build_deck_lookup_data()
        try:
            save_catalog_js(load_catalog())
        except Exception:
            pass
